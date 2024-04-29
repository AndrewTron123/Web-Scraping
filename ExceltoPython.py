import openpyxl as xl

wb = xl.load_workbook('example.xlsx')

sn = wb.sheetnames

print(sn)
sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(sheet1)
print(cellA1)
print(cellA1.value)



print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

print(sheet1.cell(1,2).value)

print(sheet1.max_row)
print(sheet1.max_column)


for i in range(1,sheet1.max_row + 1):
    print(sheet1.cell(i,2).value)

print(xl.utils.get_column_letter(5))
print(xl.utils.column_index_from_string('A'))

for current_row in sheet1['A1':'C3']:
    print(current_row)
    for currentcell in current_row:
        print(currentcell.coordinate, currentcell.value)

for current_row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, max_col=sheet1.max_column):
    print(current_row[0].value)
    print(current_row[1].value)
    print(current_row[2].value)