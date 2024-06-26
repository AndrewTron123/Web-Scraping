import openpyxl as xl
from openpyxl.styles import Font
wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'


wb.create_sheet(index = 1, title = 'Second Sheet')

#write content to a cell
ws['A1'] = 'Invoice'

ws['A1'].font = Font('Times New Roman', size = 24, italic= False, bold= True)
ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')

#unmerge
#ws.ummerge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8']= 'Total'
ws['A8'].font = Font(size = 24, bold = True)
ws['B8'] = '=sum(B2:B7)'

#change column width
ws.column_dimensions['A'].width = 25

wb.save('PythontoExcel.xlsx')


#read excel file produce report. write all contents of this file to second sheet in current workbook
#display grand total and average of amt sold and total


write_ws = wb['Second Sheet']
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

for row in read_ws:
    ls = [i.value for i in row]
    write_ws.append(ls)

max_row = write_ws.max_row

write_ws.cell(max_row + 2, 1).value = 'Totals:'
write_ws.cell(max_row + 2, 3).value = f"=sum(c1:c{max_row})"
write_ws.cell(max_row + 2, 4).value = f"=sum(d1:D{max_row})"

write_ws.cell(max_row + 3, 1).value = 'Averages'
write_ws.cell(max_row + 3, 3).value = f"=sum(c1:c{max_row})"
write_ws.cell(max_row + 3, 4).value = f"=sum(d1:D{max_row})"

wb.save('PythontoExcel.xlsx')







